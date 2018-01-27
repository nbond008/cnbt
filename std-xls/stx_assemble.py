import sys
import os
import math
import time

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

class Field(object):
    field_id = None
    data     = None

    def __init__(self, field_id, data = None):
        self.field_id = field_id
        self.data     = data

    def get_id(self):
        return self.field_id

    def set_data(self, data):
        self.data = data
        return None

    def data_append(self, point):
        self.data.append(point)
        return None

    def get_data(self):
        return self.data

def find_field_by_id(fields, field_id):
    for field in fields:
        if field.get_id() == field_id:
            return field

    return None

def read_xml(std_file):
    fields = []
    attempts = 0
    line = std_file.readline()
    while(line.strip() != '<fields>' and attempts < 30):
        line = std_file.readline()
        attempts += 1

    line = std_file.readline()
    attempts = 0
    while(line.strip() != '</fields>' and attempts < 30):
        start_index = line.find('id')
        end_index = line[(start_index + 4):].find('"') + start_index + 4
        field_id = line[(start_index + 4):end_index]
        fields.append(Field(field_id))
        line = std_file.readline().strip()
        attempts += 1

    attempts = 0
    while(line.strip() != '<systems>' and attempts < 30):
        line = std_file.readline()
        attempts += 1

    line = std_file.readline()
    line = std_file.readline()
    i = 0
    attempts = 0
    while(line.strip() != '</system>' and i < len(fields) and attempts < len(fields) * 4):
        if line.find('datum') == -1:
            value_index = line.find('value')
            if value_index == -1:
                fields[i].set_data('')
            else:
                start_index = line.find('>')
                end_index = line[(start_index + 1):].find('<') + start_index + 1
                fields[i].set_data(float((line[start_index + 1:end_index])))
            i += 1

        line = std_file.readline()
        attempts += 1

    return fields

def read_txt(txt_file):
    fields = []
    attempts = 0

    field_names = txt_file.readline().split(' ')
    for name in field_names:
        if len(name) > 1:
            fields.append(Field(name.strip(), []))

    for line in txt_file:
        i = 0
        for datum in line.split(','):
            try:
                fields[i].data_append(float(datum.strip()))
                i += 1
            except ValueError:
                fields[i].data_append(float('nan'))

    return fields

def assemble(is_cvn, fn_std, fn_bb, fn_bs, fn_ss, fn_wb, n1, n2, num_frames = 1000, static = False, sed = False):
    std_file = open(fn_std, 'r')
    bb_file  = open(fn_bb, 'r')
    bs_file  = open(fn_bs, 'r')
    ss_file  = open(fn_ss, 'r')

    print('Reading %s...' % fn_std)
    std_fields = read_xml(std_file)

    print('Reading %s,\n%s,\nand %s...' % (fn_bb, fn_bs, fn_ss))
    txt_data = [read_txt(bb_file), read_txt(bs_file), read_txt(ss_file)]

    print('Checking special characters...')
    delta = 'Delta'
    phi   = 'phi'
    chi   = 'chi'
    try:
        delta = unichr(916)
    except UnicodeEncodeError:
        print('\tUnicode character 916 (uppercase delta) not found!')
    try:
        phi = unichr(966)
    except UnicodeEncodeError:
        print('\tUnicode character 966 (lowercase phi) not found!')
    try:
        chi = unichr(967)
    except UnicodeEncodeError:
        print('\tUnicode character 967 (lowercase chi) not found!')

    print('Creating workbook...')
    wb = Workbook()
    temps = range(295, 330, 5)
    book_array = [{} for x in range(len(temps))]
    raw_array = [{} for x in range(3)]

    sheets = ['BB raw data', 'BS raw data', 'SS raw data']
    if static or sed:
        for i in range(3):
            fields  = txt_data[i]

            curraw = raw_array[i]
            for j in range(len(fields)):
                for k in range(len(fields[j].get_data())):
                    if static or sed:
                        curraw['%s%d' % (get_column_letter(j + 1), k + 2)] = fields[j].get_data()[k]

    eval_array = [{} for x in range(len(temps))]

    Z = [
        find_field_by_id(std_fields, 'Zbb').get_data(),
        find_field_by_id(std_fields, 'Zbs').get_data(),
        find_field_by_id(std_fields, 'Zsb').get_data(),
        find_field_by_id(std_fields, 'Zss').get_data()
    ]

    for t in range(len(temps)):
        temp = temps[t]
        current = book_array[t]
        cureval = eval_array[t]

        current['A1'] = 'T (K)'
        current['A2'] = temp
        current['B1'] = '%s' % chi
        current['D1'] = 'R (kcal/Kmol)'
        current['D2'] = 0.00198588

        current['F1'] = 'Zbb'
        current['G1'] = 'Zbs'
        current['H1'] = 'Zsb'
        current['I1'] = 'Zss'

        current['F2'] = Z[0]
        current['G2'] = Z[1]
        current['H2'] = Z[2]
        current['I2'] = Z[3]

        if static:
            for key in current:
                cureval[key] = current[key]

    if is_cvn:
        print 'Using CVN...'

        for t in range(len(temps)):
            temp = temps[t]
            current = book_array[t]
            cureval = eval_array[t]

            current['B2'] = '=Q2/($D$2*$A$2)'
            current['K1'] = '%sEmix/Vp' % delta
            current['K2'] = '=(G2*AI3+H2*AI3-(F2*Y3+I2*AQ3))/2'
            current['L1'] = '%sb' % phi
            current['L2'] = 0.5
            current['M1'] = '%ss' % phi
            current['M2'] = '=1-L2'
            current['N1'] = '<CVb>'
            current['N2'] = '=AVERAGE(AC3:AC1002)/%d' % n1
            current['O1'] = '<CVs>'
            current['O2'] = '=AVERAGE(AD3:AD1002)/%d' % n2
            current['P1'] = 'CVref'
            current['P2'] = '=(L2*N2)+(M2*O2)'
            current['Q1'] = '%sEmix' % delta
            current['Q2'] = '=K2*P2'

            for col in ['S', 'AA', 'AK']:
                current['%s2' % col] = 'Del_E (kcal/mol)'
            for col in ['T', 'AB', 'AL']:
                current['%s2' % col] = 'CV_Pair (A^3)'

            current['AC2'] = 'CV_Frag_1 (A^3)'
            current['AD2'] = 'CV_Frag_2 (A^3)'

            for col in ['U', 'AE', 'AM']:
                current['%s2' % col] = '%sE/Vp' % delta
            for col in ['V', 'AF', 'AN']:
                current['%s2' % col] = 'Btz Factor'
            for col in ['W', 'AG', 'AO']:
                current['%s2' % col] = 'Sum Btz'
            for col in ['X', 'AH', 'AP']:
                current['%s2' % col] = '%sE/Vp weighted' % delta
            for col in ['Y', 'AI', 'AQ']:
                current['%s2' % col] = '<%sE/Vp>' % delta

            for i in range(num_frames):
                current['S%d' % (i + 3)] = '=\'BB raw data\'!D%d' % (i + 2)
                current['T%d' % (i + 3)] = '=\'BB raw data\'!E%d' % (i + 2)
                current['U%d' % (i + 3)] = '=S%d/T%d' % (i + 3, i + 3)
                current['V%d' % (i + 3)] = '=EXP(-S%d/($A$2*$D$2))' % (i + 3)
                current['W%d' % (i + 3)] = '=W%d' % (i + 2)
                current['X%d' % (i + 3)] = '=U%d*V%d/W%d' % (i + 3, i + 3, i + 3)

                current['AA%d' % (i + 3)] = '=\'BS raw data\'!D%d' % (i + 2)
                current['AB%d' % (i + 3)] = '=\'BS raw data\'!E%d' % (i + 2)
                current['AC%d' % (i + 3)] = '=\'BS raw data\'!F%d' % (i + 2)
                current['AD%d' % (i + 3)] = '=\'BS raw data\'!G%d' % (i + 2)
                current['AE%d' % (i + 3)] = '=AA%d/AB%d' % (i + 3, i + 3)
                current['AF%d' % (i + 3)] = '=EXP(-AA%d/($A$2*$D$2))' % (i + 3)
                current['AG%d' % (i + 3)] = '=AG%d' % (i + 2)
                current['AH%d' % (i + 3)] = '=AE%d*AF%d/AG%d' % (i + 3, i + 3, i + 3)

                current['AK%d' % (i + 3)] = '=\'SS raw data\'!D%d' % (i + 2)
                current['AL%d' % (i + 3)] = '=\'SS raw data\'!E%d' % (i + 2)
                current['AM%d' % (i + 3)] = '=AK%d/AL%d' % (i + 3, i + 3)
                current['AN%d' % (i + 3)] = '=EXP(-AK%d/($A$2*$D$2))' % (i + 3)
                current['AO%d' % (i + 3)] = '=AO%d' % (i + 2)
                current['AP%d' % (i + 3)] = '=AM%d*AN%d/AO%d' % (i + 3, i + 3, i + 3)

            current['W3'] = '=SUM(V3:V1002)'
            current['AG3'] = '=SUM(AF3:AF1002)'
            current['AO3'] = '=SUM(AN3:AN1002)'

            current['Y3'] = '=SUM(X3:X1002)'
            current['AI3'] = '=SUM(AH3:AH1002)'
            current['AQ3'] = '=SUM(AP3:AP1002)'

            current['A4'] = '=B2'
            current['B4'] = '=Q2'
            current['C4'] = '=K2'
            current['D4'] = '=Y3'
            current['E4'] = '=AI3'
            current['F4'] = '=AQ3'
            current['G4'] = '=AVERAGE(T3:T1002)'
            current['H4'] = '=AVERAGE(AB3:AB1002)'
            current['I4'] = '=AVERAGE(AL3:AL1002)'
            current['J4'] = '=F2'
            current['K4'] = '=G2'
            current['L4'] = '=H2'
            current['M4'] = '=I2'

            if static:
                for key in current:
                    cureval[key] = current[key]

                cureval['W3'] = 0
                cureval['AG3'] = 0
                cureval['AO3'] = 0

                cureval['Y3'] = 0
                cureval['AI3'] = 0
                cureval['AQ3'] = 0

                cureval['N2'] = 0
                cureval['O2'] = 0

                # sums and other loops
                for row in range(3, num_frames + 3):
                    cureval['U%d' % row] = raw_array[0]['D%d' % (row - 1)]/raw_array[0]['E%d' % (row - 1)]
                    cureval['V%d' % row] = math.exp(-raw_array[0]['D%d' % (row - 1)] / (temp * 0.00198588))

                    cureval['AE%d' % row] = raw_array[1]['D%d' % (row - 1)]/raw_array[0]['E%d' % (row - 1)]
                    cureval['AF%d' % row] = math.exp(-raw_array[1]['D%d' % (row - 1)] / (temp * 0.00198588))

                    cureval['AM%d' % row] = raw_array[2]['D%d' % (row - 1)]/raw_array[0]['E%d' % (row - 1)]
                    cureval['AN%d' % row] = math.exp(-raw_array[2]['D%d' % (row - 1)] / (temp * 0.00198588))

                    cureval['N2'] += raw_array[1]['F%d' % (row - 1)] / (n1 * num_frames)
                    cureval['O2'] += raw_array[1]['G%d' % (row - 1)] / (n2 * num_frames)

                # second loop step
                for row in range(3, num_frames + 3):
                    cureval['W3'] += cureval['V%d' % row]
                    cureval['AG3'] += cureval['AF%d' % row]
                    cureval['AO3'] += cureval['AN%d' % row]

                # third loop step
                for row in range(3, num_frames + 3):
                    cureval['X%d' % row] = cureval['U%d' % row] * cureval['V%d' % row] / cureval['W3']
                    cureval['AH%d' % row] = cureval['AE%d' % row] * cureval['AF%d' % row] / cureval['AG3']
                    cureval['AP%d' % row] = cureval['AM%d' % row] * cureval['AN%d' % row] / cureval['AO3']

                # fourth loop step
                for row in range(3, num_frames + 3):
                    cureval['Y3'] += cureval['X%d' % row]
                    cureval['AI3'] += cureval['AH%d' % row]
                    cureval['AQ3'] += cureval['AP%d' % row]

                cureval['K2'] = (current['G2'] * cureval['AI3'] + current['H2'] * cureval['AI3'] - (current['F2'] * cureval['Y3'] + current['I2'] * cureval['AQ3']))/2
                cureval['M2'] = 1 - current['L2']
                cureval['P2'] = cureval['L2'] * cureval['N2'] + cureval['M2'] * cureval['O2']
                cureval['Q2'] = cureval['K2'] * cureval['P2']
                cureval['B2'] = cureval['Q2'] / (current['D2'] * current['A2'])

    else:
        print 'Not using CVN...'

        for t in range(len(temps)):
            temp = temps[t]
            current = book_array[t]
            cureval = eval_array[t]

            current['B2'] = '=K2/($D$2*$A$2)'
            current['K1'] = '%sEmix' % delta
            current['K2'] = '=(G2*W3+H2*W3-(F2*Q3+I2*AC3))/2'

            for col in ['M', 'S', 'Y']:
                current['%s2' % col] = 'Del_E (kcal/mol)'
            for col in ['N', 'T', 'Z']:
                current['%s2' % col] = 'Btz Factor'
            for col in ['O', 'U', 'AA']:
                current['%s2' % col] = 'Sum Btz'
            for col in ['P', 'V', 'AB']:
                current['%s2' % col] = '%sE/Vp weighted' % delta
            for col in ['Q', 'W', 'AC']:
                current['%s2' % col] = '<%sE/Vp>' % delta

            for i in range(num_frames):
                current['M%d' % (i + 3)] = '=\'BB raw data\'!D%d' % (i + 2)
                current['N%d' % (i + 3)] = '=EXP(-M%d/($A$2*$D$2))' % (i + 3)
                current['O%d' % (i + 3)] = '=O%d' % (i + 2)
                current['P%d' % (i + 3)] = '=M%d*N%d/O%d' % (i + 3, i + 3, i + 3)

                current['S%d' % (i + 3)] = '=\'BS raw data\'!D%d' % (i + 2)
                current['T%d' % (i + 3)] = '=EXP(-S%d/($A$2*$D$2))' % (i + 3)
                current['U%d' % (i + 3)] = '=U%d' % (i + 2)
                current['V%d' % (i + 3)] = '=S%d*T%d/U%d' % (i + 3, i + 3, i + 3)

                current['Y%d' % (i + 3)] = '=\'SS raw data\'!D%d' % (i + 2)
                current['Z%d' % (i + 3)] = '=EXP(-Y%d/($A$2*$D$2))' % (i + 3)
                current['AA%d' % (i + 3)] = '=AA%d' % (i + 2)
                current['AB%d' % (i + 3)] = '=Y%d*Z%d/AA%d' % (i + 3, i + 3, i + 3)

            current['O3'] = '=SUM(N3:N1002)'
            current['U3'] = '=SUM(T3:T1002)'
            current['AA3'] = '=SUM(Z3:Z1002)'

            current['Q3'] = '=SUM(P3:P1002)'
            current['W3'] = '=SUM(V3:V1002)'
            current['AC3'] = '=SUM(AB3:AB1002)'

            if static:
                for key in current:
                    cureval[key] = current[key]

                cureval['O3'] = 0
                cureval['U3'] = 0
                cureval['AA3'] = 0

                cureval['Q3'] = 0
                cureval['W3'] = 0
                cureval['AC3'] = 0

                # sums and other loops
                for row in range(3, num_frames + 3):
                    cureval['M%d' % row] = raw_array[0]['D%d' % (row - 1)]

                    cureval['S%d' % row] = raw_array[1]['D%d' % (row - 1)]

                    cureval['Y%d' % row] = raw_array[2]['D%d' % (row - 1)]

                    cureval['N%d' % row] = math.exp(-raw_array[0]['D%d' % (row - 1)] / (temp * 0.00198588))

                    cureval['T%d' % row] = math.exp(-raw_array[1]['D%d' % (row - 1)] / (temp * 0.00198588))

                    cureval['Z%d' % row] = math.exp(-raw_array[2]['D%d' % (row - 1)] / (temp * 0.00198588))

                # second loop step
                for row in range(3, num_frames + 3):
                    cureval['O3'] += cureval['N%d' % row]
                    cureval['U3'] += cureval['T%d' % row]
                    cureval['AA3'] += cureval['Z%d' % row]

                # third loop step
                for row in range(3, num_frames + 3):
                    cureval['P%d' % row] = cureval['M%d' % row] * cureval['N%d' % row] / cureval['O3']
                    cureval['V%d' % row] = cureval['S%d' % row] * cureval['T%d' % row] / cureval['U3']
                    cureval['AB%d' % row] = cureval['Y%d' % row] * cureval['Z%d' % row] / cureval['AA3']

                # fourth loop step
                for row in range(3, num_frames + 3):
                    cureval['Q3'] += cureval['P%d' % row]
                    cureval['W3'] += cureval['V%d' % row]
                    cureval['AC3'] += cureval['AB%d' % row]

                cureval['K2'] = (current['G2'] * cureval['W3'] + current['H2'] * cureval['W3'] - (current['F2'] * cureval['Q3'] + current['I2'] * cureval['AC3']))/2
                cureval['B2'] = cureval['K2'] / (current['D2'] * current['A2'])

    for t in range(len(temps)):
        temp = temps[t]
        current  = book_array[t]
        wb.create_sheet('%d K' % temp)
        cursheet = wb.get_sheet_by_name('%d K' % temp)

        for key in current:
            cursheet[key] = current[key]

        if is_cvn:
            cursheet.merge_cells('S1:Y1')
            cursheet.merge_cells('AA1:AI1')
            cursheet.merge_cells('AK1:AQ1')
            cursheet['S1'] = 'BB'
            cursheet['AA1'] = 'BS'
            cursheet['AK1'] = 'SS'
        else:
            cursheet.merge_cells('M1:Q1')
            cursheet.merge_cells('S1:W1')
            cursheet.merge_cells('Y1:AC1')
            cursheet['M1'] = 'BB'
            cursheet['S1'] = 'BS'
            cursheet['Y1'] = 'SS'

    for i in range(3):
        wb.create_sheet(sheets[i])
        cursheet = wb.get_sheet_by_name(sheets[i])
        fields  = txt_data[i]

        for j in range(len(fields)):
            cursheet['%s1' % get_column_letter(j + 1)] = fields[j].get_id()
            for k in range(len(fields[j].get_data())):
                cursheet['%s%d' % (get_column_letter(j + 1), k + 2)] = fields[j].get_data()[k]

    wb.create_sheet('chi vs. T')
    cursheet = wb.get_sheet_by_name('chi vs. T')

    cursheet['A1'] = 'T (K)'
    cursheet['B1'] = '%s' % chi
    for i in range(len(temps)):
        cursheet['A%d' % (i + 2)] = ('=\'%d K\'!A2' % temps[i])
        cursheet['B%d' % (i + 2)] = ('=\'%d K\'!B2' % temps[i])

    try:
        print('Graphing %s vs T...' % chi)
    except UnicodeEncodeError:
        print('Graphing chi vs T...')

    chi_graph = ScatterChart()
    chi_graph.title = ''
    chi_graph.x_axis.title = 'T (K)'
    chi_graph.y_axis.title = '%s (dimensionless)' % chi

    xvalues = Reference(cursheet, min_col = 1, min_row = 2, max_row = len(temps) + 1)
    yvalues = Reference(cursheet, min_col = 2, min_row = 2, max_row = len(temps) + 1)
    series  = Series(yvalues, xvalues, title_from_data = False)
    chi_graph.series.append(series)

    cursheet.add_chart(chi_graph, 'E3')

    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    print 'Saving workbook to %s' % fn_wb
    wb.save(fn_wb)

    if static:
        return (eval_array, raw_array)
    elif sed:
        if sys.platform == 'win32':
            os.startfile(os.path.normpath(fn_wb))
        else:
            os.system('open \'%s\'' % fn_wb)
        return (None, raw_array)
    else:
        if sys.platform == 'win32':
            os.startfile(os.path.normpath(fn_wb))
        else:
            os.system('open \'%s\'' % fn_wb)
        return None

def link(wb_base, sheet, col, row):
    return '\'[%s]%s\'!$%s$%s' % (wb_base, sheet, col, row)

def compile_dynamic(is_cvn, wb_paths, entries, fn_mds, pathchar, num_frames, sed = False):


    print 'Reading from'
    for fn_wb in wb_paths:
        print fn_wb

    print 'Waiting for files to open...\n(Nick apologizes for the delay.)'
    timeout = 5
    time.sleep(timeout)

    print('Checking special characters...')
    delta = 'Delta'
    phi   = 'phi'
    chi   = 'X'
    try:
        delta = unichr(916)
    except UnicodeEncodeError:
        print('\tUnicode character 916 (uppercase delta) not found!')
    try:
        phi = unichr(966)
    except UnicodeEncodeError:
        print('\tUnicode character 966 (lowercase phi) not found!')
    try:
        chi = unichr(967)
    except UnicodeEncodeError:
        print('\tUnicode character 967 (lowercase chi) not found!')

    print('Creating workbook...')
    mds = Workbook()
    mds.create_sheet('Data (295 K)')
    current = mds.get_sheet_by_name('Data (295 K)')

    if is_cvn:
        print('Using CVN...')

        current['A1'] = 'Run #'
        current['B1'] = '%s' % chi
        current['C1'] = '%sEmix' % delta
        current['D1'] = '%sEmix/V' % delta
        current['E1'] = 'Vref'
        current['F1'] = 'Ebb/Vbb'
        current['G1'] = 'Ebs/Vbs'
        current['H1'] = 'Ess/Vss'
        current['I1'] = 'Vbb'
        current['J1'] = 'Vbs'
        current['K1'] = 'Vss'
        current['L1'] = 'Zbb'
        current['M1'] = 'Zbs'
        current['N1'] = 'Zsb'
        current['O1'] = 'Zss'

        for i in range(0, len(wb_paths)):

            #wb = load_workbook(filename = wb_paths[i])
            #data = wb.get_sheet_by_name('295 K')

            wb_base = wb_paths[i].split(pathchar)[-1]

            current['A%d' % (i + 2)] = i + 1
            current['B%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'B', '2') #data['B2'].value
            current['C%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'Q', '2') #data['Q2'].value
            current['D%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'K', '2') #data['K2'].value
            current['E%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'P', '2') #data['P2'].value
            current['F%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'Y', '3') #data['Y3'].value
            current['G%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'AI', '3') #data['AI3'].value
            current['H%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'AQ', '3') #data['AQ3'].value

            current['I%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'G', '4') #data['Y3'].value
            current['J%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'H', '4') #data['AI3'].value
            current['K%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'I', '4') #data['AQ3'].value

            current['L%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'F', '2') #data['F2'].value
            current['M%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'G', '2') #data['G2'].value
            current['N%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'H', '2') #data['H2'].value
            current['O%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'I', '2') #data['I2'].value

    else:
        print('Not using CVN...')

        current['A1'] = 'Run #'
        current['B1'] = '%s' % chi
        current['C1'] = '%sEmix' % delta
        current['D1'] = 'Ebb'
        current['E1'] = 'Ebs'
        current['F1'] = 'Ess'
        current['G1'] = 'Zbb'
        current['H1'] = 'Zbs'
        current['I1'] = 'Zsb'
        current['J1'] = 'Zss'

        for i in range(0, len(wb_paths)):
            wb_base = wb_paths[i].split(pathchar)[-1]

            current['A%d' % (i + 2)] = '%d' % (i + 1)
            current['B%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'B', '2') #data['B2'].value
            current['C%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'K', '2') #data['K2'].value
            current['D%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'Q', '3') #data['Q3'].value
            current['E%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'W', '3') #data['W3'].value
            current['F%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'AC', '3') #data['AC3'].value
            current['G%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'F', '2') #data['F2'].value
            current['H%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'G', '2') #data['G2'].value
            current['I%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'H', '2') #data['H2'].value
            current['J%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'I', '2') #data['I2'].value

    mds.create_sheet('chi vs. T')
    current = mds.get_sheet_by_name('chi vs. T')
    current['A1'] = 'T (K)'
    temps = range(295, 330, 5)
    for t in range(0, len(temps)):
        current['A%d' % (t + 2)] = temps[t]

    chi_graph = ScatterChart()
    chi_graph.title = ''
    chi_graph.x_axis.title = 'T (K)'
    chi_graph.y_axis.title = '%s (dimensionless)' % chi

    for i in range(0, len(wb_paths)):
        current[get_column_letter(i + 2) + '1'] = 'Run %d' % (i + 1)

        wb_base = wb_paths[i].split(pathchar)[-1]
        for t in range(0, len(temps)):
            current[get_column_letter(i + 2) + str(t + 2)] = '=%s' % link(wb_base, 'chi vs. T', 'B', str(t + 2)) #data['B%d' % (i + 2)].value

        xvalues = Reference(current, min_col = 1, min_row = 2, max_row = len(temps) + 1)
        yvalues = Reference(current, min_col = i + 2, min_row = 1, max_row = len(temps) + 1)
        series  = Series(yvalues, xvalues, title_from_data = True)
        chi_graph.series.append(series)

    current.add_chart(chi_graph, 'E3')

    if sed and entries:
        sed_dict = compile_sed(entries, num_frames)
        for sheet in sed_dict:
            mds.create_sheet(sheet)
            current = mds.get_sheet_by_name(sheet)

            for key in sed_dict[sheet]:
                current[key] = sed_dict[sheet][key]

            sed_graph = ScatterChart()
            sed_graph.title = ''
            sed_graph.x_axis.title = 'Frame Index'
            sed_graph.y_axis.title = 'Energy (arbitrary)'

            for i in range(len(entries)):
                xvalues = Reference(current, min_col = 1, min_row = 2, max_row = num_frames + 1)
                yvalues = Reference(current, min_col = i + 2, min_row = 1, max_row = num_frames + 1)
                series  = Series(yvalues, xvalues, title_from_data = True)
                sed_graph.series.append(series)

            current.add_chart(sed_graph, 'L3')

    mds.remove_sheet(mds.get_sheet_by_name('Sheet'))
    print 'Saving master data sheet to %s' % fn_mds
    mds.save(fn_mds)

    if sys.platform == 'win32':
        os.startfile(os.path.normpath(fn_mds))
    else:
        os.system('open \'%s\'' % fn_mds)

def compile_static(is_cvn, entries, fn_mds, pathchar, num_frames, sed = False):
    print 'Reading from static entries'

    print('Checking special characters...')
    delta = 'Delta'
    phi   = 'phi'
    chi   = 'X'
    try:
        delta = unichr(916)
    except UnicodeEncodeError:
        print('\tUnicode character 916 (uppercase delta) not found!')
    try:
        phi = unichr(966)
    except UnicodeEncodeError:
        print('\tUnicode character 966 (lowercase phi) not found!')
    try:
        chi = unichr(967)
    except UnicodeEncodeError:
        print('\tUnicode character 967 (lowercase chi) not found!')

    print('Creating workbook...')
    mds = Workbook()
    mds.create_sheet('Data (295 K)')
    current = mds.get_sheet_by_name('Data (295 K)')

    if is_cvn:
        print('Using CVN...')

        current['A1'] = 'Run #'
        current['B1'] = '%s' % chi
        current['C1'] = '%sEmix' % delta
        current['D1'] = '%sEmix/V' % delta
        current['E1'] = 'Vref'
        current['F1'] = 'Ebb/Vbb'
        current['G1'] = 'Ebs/Vbs'
        current['H1'] = 'Ess/Vss'
        current['I1'] = 'Vbb'
        current['J1'] = 'Vbs'
        current['K1'] = 'Vss'
        current['L1'] = 'Zbb'
        current['M1'] = 'Zbs'
        current['N1'] = 'Zsb'
        current['O1'] = 'Zss'

        for i in range(len(entries)):
            entry = entries[i][0]

            current['A%d' % (i + 2)] = i + 1
            current['B%d' % (i + 2)] = entry[0]['B2']
            current['C%d' % (i + 2)] = entry[0]['Q2']
            current['D%d' % (i + 2)] = entry[0]['K2']
            current['E%d' % (i + 2)] = entry[0]['P2']
            current['F%d' % (i + 2)] = entry[0]['Y3']
            current['G%d' % (i + 2)] = entry[0]['AI3']
            current['H%d' % (i + 2)] = entry[0]['AQ3']

            Vbb = 0
            Vbs = 0
            Vss = 0

            raw_array = entries[i][1]

            for l in range(num_frames):
                Vbb += raw_array[0]['E%d' % (l + 2)] / num_frames
                Vbs += raw_array[1]['E%d' % (l + 2)] / num_frames
                Vss += raw_array[2]['E%d' % (l + 2)] / num_frames

            current['I%d' % (i + 2)] = Vbb
            current['J%d' % (i + 2)] = Vbs
            current['K%d' % (i + 2)] = Vss

            current['L%d' % (i + 2)] = entry[0]['F2']
            current['M%d' % (i + 2)] = entry[0]['G2']
            current['N%d' % (i + 2)] = entry[0]['H2']
            current['O%d' % (i + 2)] = entry[0]['I2']

    else:
        print('Not using CVN...')

        current['A1'] = 'Run #'
        current['B1'] = '%s' % chi
        current['C1'] = '%sEmix' % delta
        current['D1'] = 'Ebb'
        current['E1'] = 'Ebs'
        current['F1'] = 'Ess'
        current['G1'] = 'Zbb'
        current['H1'] = 'Zbs'
        current['I1'] = 'Zsb'
        current['J1'] = 'Zss'

        for i in range(0, len(entries)):
            entry = entries[i][0]

            current['A%d' % (i + 2)] = '%d' % (i + 1)
            current['B%d' % (i + 2)] = entry[0]['B2']
            current['C%d' % (i + 2)] = entry[0]['K2']
            current['D%d' % (i + 2)] = entry[0]['Q3']
            current['E%d' % (i + 2)] = entry[0]['W3']
            current['F%d' % (i + 2)] = entry[0]['AC3']
            current['G%d' % (i + 2)] = entry[0]['F2']
            current['H%d' % (i + 2)] = entry[0]['G2']
            current['I%d' % (i + 2)] = entry[0]['H2']
            current['J%d' % (i + 2)] = entry[0]['I2']

    mds.create_sheet('chi vs. T')
    current = mds.get_sheet_by_name('chi vs. T')

    current['A1'] = 'T (K)'
    temps = range(295, 330, 5)
    for t in range(0, len(temps)):
        current['A%d' % (t + 2)] = temps[t]

    chi_graph = ScatterChart()
    chi_graph.title = ''
    chi_graph.x_axis.title = 'T (K)'
    chi_graph.y_axis.title = '%s (dimensionless)' % chi

    for i in range(len(entries)):
        entry = entries[i][0]

        current[get_column_letter(i + 2) + '1'] = 'Run %d' % (i + 1)

        for t in range(len(temps)):
            current[get_column_letter(i + 2) + str(t + 2)] = entry[t]['B2']

        xvalues = Reference(current, min_col = 1, min_row = 2, max_row = len(temps) + 1)
        yvalues = Reference(current, min_col = i + 2, min_row = 1, max_row = len(temps) + 1)
        series  = Series(yvalues, xvalues, title_from_data = True)
        chi_graph.series.append(series)

    current.add_chart(chi_graph, 'E3')

    if sed:
        sed_dict = compile_sed(entries, num_frames)
        for sheet in sed_dict:
            mds.create_sheet(sheet)
            current = mds.get_sheet_by_name(sheet)

            for key in sed_dict[sheet]:
                current[key] = sed_dict[sheet][key]

            sed_graph = ScatterChart()
            sed_graph.title = ''
            sed_graph.x_axis.title = 'Frame Index'
            sed_graph.y_axis.title = 'Energy (arbitrary)'

            for i in range(len(entries)):
                xvalues = Reference(current, min_col = 1, min_row = 2, max_row = num_frames + 1)
                yvalues = Reference(current, min_col = i + 2, min_row = 1, max_row = num_frames + 1)
                series  = Series(yvalues, xvalues, title_from_data = True)
                sed_graph.series.append(series)

            current.add_chart(sed_graph, 'L3')

    mds.remove_sheet(mds.get_sheet_by_name('Sheet'))
    print 'Saving master data sheet to %s' % fn_mds
    mds.save(fn_mds)

    if sys.platform == 'win32':
        os.startfile(os.path.normpath(fn_mds))
    else:
        os.system('open \'%s\'' % fn_mds)

def compile_sed(entries, num_frames):
    sed_dict = {
        'SED BB' : {},
        'SED BS' : {},
        'SED SS' : {}
    }

    for key in sed_dict:
        sed_dict[key]['A1'] = 'Frame Index'
        for i in range(1, num_frames + 1):
            sed_dict[key]['A%d' % (i + 1)] = i
        for i in range(len(entries) + 1):
            sed_dict[key]['%s1' % get_column_letter(i + 1)] = 'Run %d' % i

    for j in range(len(entries)):
        bb = [0 for x in range(num_frames)]
        bs = [0 for x in range(num_frames)]
        ss = [0 for x in range(num_frames)]

        for i in range(num_frames):
            bb[i] = entries[j][1][0]['D%d' % (i + 2)]
            bs[i] = entries[j][1][1]['D%d' % (i + 2)]
            ss[i] = entries[j][1][2]['D%d' % (i + 2)]

        bb.sort()
        bs.sort()
        ss.sort()

        for i in range(num_frames):
            sed_dict['SED BB']['%s%d' % (get_column_letter(j + 2), i + 2)] = bb[i]
            sed_dict['SED BS']['%s%d' % (get_column_letter(j + 2), i + 2)] = bs[i]
            sed_dict['SED SS']['%s%d' % (get_column_letter(j + 2), i + 2)] = ss[i]

    return sed_dict
