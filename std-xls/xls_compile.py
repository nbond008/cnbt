from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
#import re
import os
from sys import platform
import time

def link(wb_base, sheet, col, row):
    #get cell
    #print '=\'%s[%s]%s\'!$%s$%s' % ('', wb_base, sheet, col, row)
    return '\'[%s]%s\'!$%s$%s' % (wb_base, sheet, col, row)

def main(is_cvn, wb_paths, fn_mds, pathchar):

    print 'Reading from'
    for fn_wb in wb_paths:
        print fn_wb

    print 'Waiting for files to open...\n(Nick apologizes for the delay.)'
    timeout = 5
    time.sleep(timeout)

    #print fn_mds

    print('Checking special characters...')
    delta = 'Delta'
    phi   = 'phi'
    chi   = 'X'
    try:
        delta = unichr(916)
    except UnicodeEncodeError:
        print('\tUnicode character 916 (uppercase delta) not found!')
    try:
        phi = unichr(966)
    except UnicodeEncodeError:
        print('\tUnicode character 966 (lowercase phi) not found!')
    try:
        chi = unichr(967)
    except UnicodeEncodeError:
        print('\tUnicode character 967 (lowercase chi) not found!')

    print('Creating workbook...')
    mds = Workbook()
    mds.create_sheet('Data (295 K)')
    current = mds.get_sheet_by_name('Data (295 K)')

    if is_cvn:
        print('Using CVN...')

        current['A1'] = 'Run #'
        current['B1'] = '%s' % chi
        current['C1'] = '%sEmix' % delta
        current['D1'] = '%sEmix/V' % delta
        current['E1'] = 'Vref'
        current['F1'] = 'Ebb/Vbb'
        current['G1'] = 'Ebs/Vbs'
        current['H1'] = 'Ess/Vss'
        current['I1'] = 'Vbb'
        current['J1'] = 'Vbs'
        current['K1'] = 'Vss'
        current['L1'] = 'Zbb'
        current['M1'] = 'Zbs'
        current['N1'] = 'Zsb'
        current['O1'] = 'Zss'

        for i in range(0, len(wb_paths)):

            #wb = load_workbook(filename = wb_paths[i])
            #data = wb.get_sheet_by_name('295 K')

            wb_base = wb_paths[i].split(pathchar)[-1]

            current['A%d' % (i + 2)] = i + 1
            current['B%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'B', '2') #data['B2'].value
            current['C%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'Q', '2') #data['Q2'].value
            current['D%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'K', '2') #data['K2'].value
            current['E%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'P', '2') #data['P2'].value
            current['F%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'Y', '3') #data['Y3'].value
            current['G%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'AI', '3') #data['AI3'].value
            current['H%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'AQ', '3') #data['AQ3'].value

            current['I%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'G', '4') #data['Y3'].value
            current['J%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'H', '4') #data['AI3'].value
            current['K%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'I', '4') #data['AQ3'].value

            #current['I%d' % (i + 2)] = '=SUM(%s)' % link(wb_base, '295 K', 'AQ', '3:1002') #data['AQ3'].value
            #current['J%d' % (i + 2)] = '=SUM(%s)' % link(wb_base, '295 K', 'AQ', '3:1002') #data['AQ3'].value
            #current['K%d' % (i + 2)] = '=SUM(%s)' % link(wb_base, '295 K', 'AQ', '3:1002') #data['AQ3'].value

            current['L%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'F', '2') #data['F2'].value
            current['M%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'G', '2') #data['G2'].value
            current['N%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'H', '2') #data['H2'].value
            current['O%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'I', '2') #data['I2'].value

    else:
        print('Not using CVN...')

        current['A1'] = 'Run #'
        current['B1'] = '%s' % chi
        current['C1'] = '%sEmix' % delta
        current['D1'] = 'Ebb'
        current['E1'] = 'Ebs'
        current['F1'] = 'Ess'
        current['G1'] = 'Zbb'
        current['H1'] = 'Zbs'
        current['I1'] = 'Zsb'
        current['J1'] = 'Zss'

        for i in range(0, len(wb_paths)):
            # wb = load_workbook(filename = wb_paths[i])
            # data = wb.get_sheet_by_name('295 K')

            wb_base = wb_paths[i].split(pathchar)[-1]

            current['A%d' % (i + 2)] = '%d' % (i + 1)
            current['B%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'B', '2') #data['B2'].value
            current['C%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'K', '2') #data['K2'].value
            current['D%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'Q', '3') #data['Q3'].value
            current['E%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'W', '3') #data['W3'].value
            current['F%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'AC', '3') #data['AC3'].value
            current['G%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'F', '2') #data['F2'].value
            current['H%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'G', '2') #data['G2'].value
            current['I%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'H', '2') #data['H2'].value
            current['J%d' % (i + 2)] = '=%s' % link(wb_base, '295 K', 'I', '2') #data['I2'].value

    mds.create_sheet('chi vs. T')
    current = mds.get_sheet_by_name('chi vs. T')
    #current = mds.worksheets[7]

    current['A1'] = 'T (K)'
    temps = range(295, 330, 5)
    for t in range(0, len(temps)):
        current['A%d' % (t + 2)] = temps[t]

    chi_graph = ScatterChart()
    chi_graph.title = ''
    chi_graph.x_axis.title = 'T (K)'
    chi_graph.y_axis.title = '%s (dimensionless)' % chi

    for i in range(0, len(wb_paths)):
        # wb = load_workbook(filename = wb_paths[i])
        # data = wb.get_sheet_by_name('chi vs. T')

        current[get_column_letter(i + 2) + '1'] = 'Run %d' % (i + 1)

        wb_base = wb_paths[i].split(pathchar)[-1]
        for t in range(0, len(temps)):
            current[get_column_letter(i + 2) + str(t + 2)] = '=%s' % link(wb_base, 'chi vs. T', 'B', str(t + 2)) #data['B%d' % (i + 2)].value

        xvalues = Reference(current, min_col = 1, min_row = 2, max_row = len(temps) + 1)
        yvalues = Reference(current, min_col = i + 2, min_row = 1, max_row = len(temps) + 1)
        series  = Series(yvalues, xvalues, title_from_data = True)
        chi_graph.series.append(series)

    current.add_chart(chi_graph, 'E3')

    mds.remove_sheet(mds.get_sheet_by_name('Sheet'))
    print 'Saving master data sheet to %s' % fn_mds
    mds.save(fn_mds)

    if platform == 'win32':
        os.startfile(os.path.normpath(fn_mds))
    else:
        os.system('open \'%s\'' % fn_mds)
