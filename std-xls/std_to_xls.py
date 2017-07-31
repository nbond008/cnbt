import sys
import os
from openpyxl import Workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

class Field(object):
    field_id = None
    data     = None

    def __init__(self, field_id, data = None):
        self.field_id = field_id
        self.data     = data

    def get_id(self):
        return self.field_id

    def set_data(self, data):
        self.data = data
        return None

    def data_append(self, point):
        self.data.append(point)
        return None

    def get_data(self):
        return self.data

def find_field_by_id(fields, field_id):
    for field in fields:
        if field.get_id() == field_id:
            return field

    return None

def read_xml(std_file):
    fields = []
    attempts = 0
    line = std_file.readline()
    while(line.strip() != '<fields>' and attempts < 30):
        line = std_file.readline()
        attempts += 1

    line = std_file.readline()
    attempts = 0
    while(line.strip() != '</fields>' and attempts < 30):
        start_index = line.find('id')
        end_index = line[(start_index + 4):].find('"') + start_index + 4
        field_id = line[(start_index + 4):end_index]
        fields.append(Field(field_id))
        line = std_file.readline().strip()
        attempts += 1

    attempts = 0
    while(line.strip() != '<systems>' and attempts < 30):
        line = std_file.readline()
        attempts += 1

    line = std_file.readline()
    line = std_file.readline()
    i = 0
    attempts = 0
    while(line.strip() != '</system>' and i < len(fields) and attempts < len(fields) * 4):
        if line.find('datum') == -1:
            value_index = line.find('value')
            if value_index == -1:
                fields[i].set_data('')
            else:
                start_index = line.find('>')
                end_index = line[(start_index + 1):].find('<') + start_index + 1
                fields[i].set_data(float((line[start_index + 1:end_index])))
            i += 1

        line = std_file.readline()
        attempts += 1

    return fields

def read_txt(txt_file):
    fields = []
    attempts = 0

    field_names = txt_file.readline().split(' ')
    for name in field_names:
        if len(name) > 1:
            fields.append(Field(name.strip(), []))

    for line in txt_file:
        i = 0
        for datum in line.split(','):
            try:
                fields[i].data_append(float(datum.strip()))
                i += 1
            except ValueError:
                fields[i].data_append(float('nan'))

    return fields

def main(is_cvn, fn_std, fn_bb, fn_bs, fn_ss, fn_wb, n1, n2):

    # print fn_wb
    # return

    std_file = open(fn_std, 'r')
    bb_file  = open(fn_bb, 'r')
    bs_file  = open(fn_bs, 'r')
    ss_file  = open(fn_ss, 'r')

    print('Reading %s...' % fn_std)
    std_fields = read_xml(std_file)

    print('Reading %s,\n%s,\nand %s...' % (fn_bb, fn_bs, fn_ss))
    txt_data = [read_txt(bb_file), read_txt(bs_file), read_txt(ss_file)]

    print('Checking special characters...')
    delta = 'Delta'
    phi   = 'phi'
    chi   = 'chi'
    try:
        delta = unichr(916)
    except UnicodeEncodeError:
        print('\tUnicode character 916 (uppercase delta) not found!')
    try:
        phi = unichr(966)
    except UnicodeEncodeError:
        print('\tUnicode character 966 (lowercase phi) not found!')
    try:
        chi = unichr(967)
    except UnicodeEncodeError:
        print('\tUnicode character 967 (lowercase chi) not found!')

    print('Creating workbook...')
    wb = Workbook()
    temps = range(295, 330, 5)
    for temp in temps:
        wb.create_sheet('%d K' % temp)
        current = wb.get_sheet_by_name('%d K' % temp)

        current['A1'] = 'T (K)'
        current['A2'] = temp
        current['B1'] = '%s' % chi
        current['D1'] = 'R (kcal/Kmol)'
        current['D2'] = 0.00198588

        current['F1'] = 'Zbb'
        current['G1'] = 'Zbs'
        current['H1'] = 'Zsb'
        current['I1'] = 'Zss'

        if temp == temps[0]:
            current['F2'] = find_field_by_id(std_fields, 'Zbb').get_data()
            current['G2'] = find_field_by_id(std_fields, 'Zbs').get_data()
            current['H2'] = find_field_by_id(std_fields, 'Zsb').get_data()
            current['I2'] = find_field_by_id(std_fields, 'Zss').get_data()
        else:
            current['F2'] = '=\'%d K\'!F2' % temps[0]
            current['G2'] = '=\'%d K\'!G2' % temps[0]
            current['H2'] = '=\'%d K\'!H2' % temps[0]
            current['I2'] = '=\'%d K\'!I2' % temps[0]

    if is_cvn:
        print 'Using CVN...'

        for temp in temps:
            current = wb.get_sheet_by_name('%d K' % temp)

            current['B2'] = '=Q2/($D$2*$A$2)'
            current['K1'] = '%sEmix/Vp' % delta
            current['K2'] = '=(G2*AI3+H2*AI3-(F2*Y3+I2*AQ3))/2'
            current['L1'] = '%sb' % phi
            current['L2'] = 0.5
            current['M1'] = '%ss' % phi
            current['M2'] = '=1-L2'
            current['N1'] = '<CVb>'
            current['N2'] = '=AVERAGE(AC3:AC1002)/%d' % n1
            current['O1'] = '<CVs>'
            current['O2'] = '=AVERAGE(AD3:AD1002)/%d' % n2
            current['P1'] = 'CVref'
            current['P2'] = '=(L2*N2)+(M2*O2)'
            current['Q1'] = '%sEmix' % delta
            current['Q2'] = '=K2*P2'

            current.merge_cells('S1:Y1')
            current.merge_cells('AA1:AI1')
            current.merge_cells('AK1:AQ1')
            current['S1'] = 'BB'
            current['AA1'] = 'BS'
            current['AK1'] = 'SS'

            for col in ['S', 'AA', 'AK']:
                current['%s2' % col] = 'Del_E (kcal/mol)'
            for col in ['T', 'AB', 'AL']:
                current['%s2' % col] = 'CV_Pair (A^3)'

            current['AC2'] = 'CV_Frag_1 (A^3)'
            current['AD2'] = 'CV_Frag_2 (A^3)'

            for col in ['U', 'AE', 'AM']:
                current['%s2' % col] = '%sE/Vp' % delta
            for col in ['V', 'AF', 'AN']:
                current['%s2' % col] = 'Btz Factor'
            for col in ['W', 'AG', 'AO']:
                current['%s2' % col] = 'Sum Btz'
            for col in ['X', 'AH', 'AP']:
                current['%s2' % col] = '%sE/Vp weighted' % delta
            for col in ['Y', 'AI', 'AQ']:
                current['%s2' % col] = '<%sE/Vp>' % delta

            for i in range(1000):
                current['S%d' % (i + 3)] = '=\'BB raw data\'!D%d' % (i + 2)
                current['T%d' % (i + 3)] = '=\'BB raw data\'!E%d' % (i + 2)
                current['U%d' % (i + 3)] = '=S%d/T%d' % (i + 3, i + 3)
                current['V%d' % (i + 3)] = '=EXP(-S%d/($A$2*$D$2))' % (i + 3)
                current['W%d' % (i + 3)] = '=W%d' % (i + 2)
                current['X%d' % (i + 3)] = '=U%d*V%d/W%d' % (i + 3, i + 3, i + 3)

                current['AA%d' % (i + 3)] = '=\'BS raw data\'!D%d' % (i + 2)
                current['AB%d' % (i + 3)] = '=\'BS raw data\'!E%d' % (i + 2)
                current['AC%d' % (i + 3)] = '=\'BS raw data\'!F%d' % (i + 2)
                current['AD%d' % (i + 3)] = '=\'BS raw data\'!G%d' % (i + 2)
                current['AE%d' % (i + 3)] = '=AA%d/AB%d' % (i + 3, i + 3)
                current['AF%d' % (i + 3)] = '=EXP(-AA%d/($A$2*$D$2))' % (i + 3)
                current['AG%d' % (i + 3)] = '=AG%d' % (i + 2)
                current['AH%d' % (i + 3)] = '=AE%d*AF%d/AG%d' % (i + 3, i + 3, i + 3)

                current['AK%d' % (i + 3)] = '=\'SS raw data\'!D%d' % (i + 2)
                current['AL%d' % (i + 3)] = '=\'SS raw data\'!E%d' % (i + 2)
                current['AM%d' % (i + 3)] = '=AK%d/AL%d' % (i + 3, i + 3)
                current['AN%d' % (i + 3)] = '=EXP(-AK%d/($A$2*$D$2))' % (i + 3)
                current['AO%d' % (i + 3)] = '=AO%d' % (i + 2)
                current['AP%d' % (i + 3)] = '=AM%d*AN%d/AO%d' % (i + 3, i + 3, i + 3)

            current['W3'] = '=SUM(V3:V1002)'
            current['AG3'] = '=SUM(AF3:AF1002)'
            current['AO3'] = '=SUM(AN3:AN1002)'

            current['Y3'] = '=SUM(X3:X1002)'
            current['AI3'] = '=SUM(AH3:AH1002)'
            current['AQ3'] = '=SUM(AP3:AP1002)'

            current['A4'] = '=B2'
            current['B4'] = '=Q2'
            current['C4'] = '=K2'
            current['D4'] = '=Y3'
            current['E4'] = '=AI3'
            current['F4'] = '=AQ3'
            current['G4'] = '=AVERAGE(T3:T1002)'
            current['H4'] = '=AVERAGE(AB3:AB1002)'
            current['I4'] = '=AVERAGE(AL3:AL1002)'
            current['J4'] = '=F2'
            current['K4'] = '=G2'
            current['L4'] = '=H2'
            current['M4'] = '=I2'

    else:
        print 'Not using CVN...'

        for temp in temps:
            current = wb.get_sheet_by_name('%d K' % temp)

            current['B2'] = '=K2/($D$2*$A$2)'
            current['K1'] = '%sEmix' % delta
            current['K2'] = '=(G2*W3+H2*W3-(F2*Q3+I2*AC3))/2'

            current.merge_cells('M1:Q1')
            current.merge_cells('S1:W1')
            current.merge_cells('Y1:AC1')
            current['M1'] = 'BB'
            current['S1'] = 'BS'
            current['Y1'] = 'SS'

            for col in ['M', 'S', 'Y']:
                current['%s2' % col] = 'Del_E (kcal/mol)'
            for col in ['N', 'T', 'Z']:
                current['%s2' % col] = 'Btz Factor'
            for col in ['O', 'U', 'AA']:
                current['%s2' % col] = 'Sum Btz'
            for col in ['P', 'V', 'AB']:
                current['%s2' % col] = '%sE/Vp weighted' % delta
            for col in ['Q', 'W', 'AC']:
                current['%s2' % col] = '<%sE/Vp>' % delta

            for i in range(1000):
                current['M%d' % (i + 3)] = '=\'BB raw data\'!D%d' % (i + 2)
                current['N%d' % (i + 3)] = '=EXP(-M%d/($A$2*$D$2))' % (i + 3)
                current['O%d' % (i + 3)] = '=O%d' % (i + 2)
                current['P%d' % (i + 3)] = '=M%d*N%d/O%d' % (i + 3, i + 3, i + 3)

                current['S%d' % (i + 3)] = '=\'BS raw data\'!D%d' % (i + 2)
                current['T%d' % (i + 3)] = '=EXP(-S%d/($A$2*$D$2))' % (i + 3)
                current['U%d' % (i + 3)] = '=U%d' % (i + 2)
                current['V%d' % (i + 3)] = '=S%d*T%d/U%d' % (i + 3, i + 3, i + 3)

                current['Y%d' % (i + 3)] = '=\'SS raw data\'!D%d' % (i + 2)
                current['Z%d' % (i + 3)] = '=EXP(-Y%d/($A$2*$D$2))' % (i + 3)
                current['AA%d' % (i + 3)] = '=AA%d' % (i + 2)
                current['AB%d' % (i + 3)] = '=Y%d*Z%d/AA%d' % (i + 3, i + 3, i + 3)

            current['O3'] = '=SUM(N3:N1002)'
            current['U3'] = '=SUM(T3:T1002)'
            current['AA3'] = '=SUM(Z3:Z1002)'

            current['Q3'] = '=SUM(P3:P1002)'
            current['W3'] = '=SUM(V3:V1002)'
            current['AC3'] = '=SUM(AB3:AB1002)'

    wb.create_sheet('chi vs. T')
    current = wb.get_sheet_by_name('chi vs. T')

    current['A1'] = 'T (K)'
    current['B1'] = '%s' % chi
    for i in range(len(temps)):
        current['A%d' % (i + 2)] = ('=\'%d K\'!A2' % temps[i])
        current['B%d' % (i + 2)] = ('=\'%d K\'!B2' % temps[i])

    try:
        print('Graphing %s vs T...' % chi)
    except UnicodeEncodeError:
        print('Graphing chi vs T...')

    chi_graph = ScatterChart()
    chi_graph.title = ''
    chi_graph.x_axis.title = 'T (K)'
    chi_graph.y_axis.title = '%s (dimensionless)' % chi

    xvalues = Reference(current, min_col = 1, min_row = 2, max_row = len(temps) + 1)
    yvalues = Reference(current, min_col = 2, min_row = 2, max_row = len(temps) + 1)
    series  = Series(yvalues, xvalues, title_from_data = False)
    chi_graph.series.append(series)

    current.add_chart(chi_graph, 'E3')

    sheets = ['BB raw data', 'BS raw data', 'SS raw data']
    for i in range(3):
        wb.create_sheet(sheets[i])
        current = wb.get_sheet_by_name(sheets[i])
        fields  = txt_data[i]
        for j in range(len(fields)):
            current['%s1' % chr(j + 65)] = fields[j].get_id()
            for k in range(len(fields[j].get_data())):
                current['%s%d' % (chr(j + 65), k + 2)] = fields[j].get_data()[k]

    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    print 'Saving workbook to %s' % fn_wb
    wb.save(fn_wb)

    if sys.platform == 'win32':
        os.startfile(os.path.normpath(fn_wb))
    else:
        os.system('open \'%s\'' % fn_wb)

if __name__ == '__main__':

    if (len(sys.argv) < 6 or len(sys.argv) > 7):
        print('Please use the format\npython %s [ -cvn ] <std>.std <bb>-out.txt <bs>-out.txt <ss>-out.txt <xls>.xlsx.' % sys.argv[0])
        sys.exit(0)

    if sys.argv[len(sys.argv) - 1][len(sys.argv[len(sys.argv) - 1]) - 5:] != '.xlsx':
        print('Please specify an output file with the extension .xlsx.')
        sys.exit(0)

    is_cvn = False
    fn_std = ''
    fn_bb  = ''
    fn_bs  = ''
    fn_ss  = ''
    fn_wb  = ''

    if sys.argv[1] != '-cvn':
        fn_std = sys.argv[1]
        fn_bb  = sys.argv[2]
        fn_bs  = sys.argv[3]
        fn_ss  = sys.argv[4]
        fn_wb  = sys.argv[5]
        #print('No CVN')
    else:
        fn_std = sys.argv[2]
        fn_bb  = sys.argv[3]
        fn_bs  = sys.argv[4]
        fn_ss  = sys.argv[5]
        fn_wb  = sys.argv[6]
        is_cvn = True
        #print('Using CVN')

    main(is_cvn, fn_std, fn_bb, fn_bs, fn_ss, fn_wb)
